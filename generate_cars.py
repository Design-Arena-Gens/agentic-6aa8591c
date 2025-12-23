import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Данные о 5 машинах США
cars_data = {
    'Марка': ['Ford', 'Chevrolet', 'Tesla', 'Jeep', 'Dodge'],
    'Модель': ['Mustang', 'Corvette', 'Model 3', 'Wrangler', 'Challenger'],
    'Год': [2023, 2024, 2023, 2022, 2023],
    'Цена ($)': [55000, 75000, 45000, 48000, 62000],
    'Мощность (л.с.)': [450, 495, 283, 285, 485],
    'Тип двигателя': ['Бензин V8', 'Бензин V8', 'Электро', 'Бензин V6', 'Бензин V8']
}

# Создаем DataFrame
df = pd.DataFrame(cars_data)

# Добавляем вычисляемые колонки
df['Цена/Мощность'] = (df['Цена ($)'] / df['Мощность (л.с.)']).round(2)
df['Категория'] = df['Цена ($)'].apply(lambda x: 'Премиум' if x > 60000 else 'Средний' if x > 50000 else 'Доступный')

# Сохраняем в Excel
excel_file = 'usa_cars.xlsx'
df.to_excel(excel_file, index=False, sheet_name='Автомобили США')

# Форматируем Excel файл
wb = load_workbook(excel_file)
ws = wb.active

# Стиль заголовков
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Автоширина колонок
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Добавляем статистику
stats_row = ws.max_row + 2
ws[f'A{stats_row}'] = 'Статистика:'
ws[f'A{stats_row}'].font = Font(bold=True, size=11)

stats_row += 1
ws[f'A{stats_row}'] = f'Средняя цена: ${df["Цена ($)"].mean():,.0f}'
stats_row += 1
ws[f'A{stats_row}'] = f'Средняя мощность: {df["Мощность (л.с.)"].mean():.0f} л.с.'
stats_row += 1
ws[f'A{stats_row}'] = f'Самая дорогая: {df.loc[df["Цена ($)"].idxmax(), "Марка"]} {df.loc[df["Цена ($)"].idxmax(), "Модель"]}'
stats_row += 1
ws[f'A{stats_row}'] = f'Самая мощная: {df.loc[df["Мощность (л.с.)"].idxmax(), "Марка"]} {df.loc[df["Мощность (л.с.)"].idxmax(), "Модель"]}'

wb.save(excel_file)

print("=" * 80)
print("РЕЗУЛЬТАТЫ ГЕНЕРАЦИИ АВТОМОБИЛЕЙ США")
print("=" * 80)
print("\nТаблица автомобилей:")
print(df.to_string(index=False))
print("\n" + "=" * 80)
print("СТАТИСТИКА:")
print("=" * 80)
print(f"Средняя цена: ${df['Цена ($)'].mean():,.0f}")
print(f"Средняя мощность: {df['Мощность (л.с.)'].mean():.0f} л.с.")
print(f"Минимальная цена: ${df['Цена ($)'].min():,} ({df.loc[df['Цена ($)'].idxmin(), 'Марка']} {df.loc[df['Цена ($)'].idxmin(), 'Модель']})")
print(f"Максимальная цена: ${df['Цена ($)'].max():,} ({df.loc[df['Цена ($)'].idxmax(), 'Марка']} {df.loc[df['Цена ($)'].idxmax(), 'Модель']})")
print(f"Самая мощная: {df.loc[df['Мощность (л.с.)'].idxmax(), 'Марка']} {df.loc[df['Мощность (л.с.)'].idxmax(), 'Модель']} - {df['Мощность (л.с.)'].max()} л.с.")
print(f"\nРаспределение по категориям:")
print(df['Категория'].value_counts().to_string())
print("\n" + "=" * 80)
print(f"✓ Excel файл '{excel_file}' успешно создан и отформатирован!")
print("=" * 80)
